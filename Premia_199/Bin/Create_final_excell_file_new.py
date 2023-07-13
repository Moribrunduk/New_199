import json
import os
from win32com.client import Dispatch

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

import configparser
import math
from create_form_new import CREATE_FORM
from Create_file_with_user_input_in_table import CREATE_FILE
from PyQt5.QtCore import QSettings


class LOAD_DATA():
    def __init__(self, proffession_number):
        self.proffession_number = str(proffession_number)
        self.SETINGS  = configparser.ConfigParser()
        self.SETINGS.read('Main\Settings_199\SETTINGS.ini', encoding="utf-8")
        self.settings = QSettings("NITIC")
        self.settings.beginGroup("199_settings")
    
    def Tarif(self):
        self.SETINGS_tarif = {self.proffession_number:
                                {
                                    3:self.SETINGS[self.proffession_number]["cv_three_tarif"],
                                    4:self.SETINGS[self.proffession_number]["cv_four_tarif"],
                                    5:self.SETINGS[self.proffession_number]["cv_five_tarif"],
                                    6:self.SETINGS[self.proffession_number]["cv_six_tarif"]}}
        return self.SETINGS_tarif
    
    def Procent(self):
        self.SETINGS_procent = {self.proffession_number : self.SETINGS[self.proffession_number]["procent_text"] }
        
        return self.SETINGS_procent
    
    def Path_with_json(self):
        self.path_with_json = self.settings.value(f"Path_with_json_{self.proffession_number}")
        return self.path_with_json
    
    def Path_for_excell(self):
        CE = CREATE_FILE(self.proffession_number)
        CE.Main()
        self.path_for_excell = self.settings.value(f"Path_for_excell_{self.proffession_number}")
        return self.path_for_excell
    
    def Current_place(self):
        self.SETINGS_current_place = self.Path_with_json().partition("\\")[-1].partition("\\")[-1].split("\\")[0]
        return self.SETINGS_current_place

    def Current_data(self):
        self.SETINGS_current_data = self.Path_with_json().partition("\\")[-1].partition("\\")[-1].partition("\\")[-1]
        return self.SETINGS_current_data
    
    def Current_month_str(self):
        return self.Current_data().split("\\")[1]
    
    def Current_month_int(self):
        self.SETINGS_current_year = self.Current_data().split("\\")[0]
        self.SETINGS_current_month = self.Current_data().split("\\")[1]
        Year = {"01":"январь",
                "02":"февраль",
                "03":"март",
                "04":"апрель",
                "05":"май",
                "06":"июнь",
                "07":"июль",
                "08":"август",
                "09":"сентябрь",
                "10":"октябрь",
                "11":"ноябрь",
                "12":"декабрь"}
        for key,vallue in Year.items():
            try:
                if vallue == self.SETINGS_current_month:
                    self.month = key
            except Exception as ex:
                print(ex)
                self.month = "00"
        self.year = self.SETINGS_current_year[2:4]
        return self.month
    
    def Current_year_int(self):
        return self.Current_data().split("\\")[0]

    def Current_year_str(self):
        self.SETINGS_current_year = self.Current_data().split("\\")[0]
     
        Year = {"01":"январь",
                "02":"февраль",
                "03":"март",
                "04":"апрель",
                "05":"май",
                "06":"июнь",
                "07":"июль",
                "08":"август",
                "09":"сентябрь",
                "10":"октябрь",
                "11":"ноябрь",
                "12":"декабрь"}
        for key,vallue in Year.items():
            try:
                if vallue == self.SETINGS_current_month:
                    self.month = key
            except Exception as ex:
                print(ex)
                self.month = "00"
        self.year = self.SETINGS_current_year[2:4]

        return self.year
    
    def Current_main_person(self):
        return self.SETINGS["Excell_data"]["current_profession_index"]
    def Current_main_person_name(self):
        return self.SETINGS["Excell_data"]["current_main_name_index"]
    def Current_botiz(self):
        return self.SETINGS["Excell_data"]["current_botiz_profession_index"]
    def Current_botiz_name(self):
        return self.SETINGS["Excell_data"]["current_botiz_name_index"]

    def Current_vedomosti(self):
        if self.proffession_number =="87100":
            if self.Current_place() == "ц.42":
                return eval(self.SETINGS["Excell_data"]["vedomosti"])["42"]
            if self.Current_place() == "КСП":
                return eval(self.SETINGS["Excell_data"]["vedomosti"])["7"]
            if self.Current_place() == "ССП Э1":
                return eval(self.SETINGS["Excell_data"]["vedomosti"])["50"]
            if self.Current_place() == "ССП Э2":
                return eval(self.SETINGS["Excell_data"]["vedomosti"])["55"]
        elif self.proffession_number =="87200":
            return eval(self.SETINGS["Excell_data"]["vedomosti"])["PZRS"]
        elif self.proffession_number =="08300":
            return eval(self.SETINGS["Excell_data"]["vedomosti"])["foto"]
        else:
            return "____"

    def All_data(self):
        with open(self.Path_with_json(), "r", encoding="utf-8") as file:
                self.all_data = json.load(file)
        
        return self.all_data

    def Subtitutes(self):
        self.substitutes = configparser.ConfigParser()
        self.substitutes.read(f"{self.Path_for_excell()}",encoding='utf-8')
        return self.substitutes

class CREATE_DATA_TO_WRITE_IN_EXCELL():
    def __init__(self, proffession_number):
        self.proffession_number = str(proffession_number)
        super(CREATE_DATA_TO_WRITE_IN_EXCELL, self).__init__()
        self.DATA = LOAD_DATA(self.proffession_number)
        
    
    def Main(self):
        return self.CreateList()

    def CreateList(self):

        """
        функция которая загружает данные из файла datalist of subtitutes.ini
        (Замещаемый, причина отсутствия, дата начала отсутствия, замещающий)
        и приводит в список (("Воронов Е.М., табельный 479",
                            "Дефектоскопист РГГ",
                            "Отпуск",
                            "7.11-31.11.2020",
                            "тариф(10988)",
                            "Илюхин Ю.В., табельный 473",
                            "Дефектоскопист РГГ")

        """
        self.all_data = self.DATA.All_data()
        self.tabels = self.all_data["шифр"][str(self.proffession_number)]["Табельный"]

        # задаем название професии
        profession_name = ""
        if str(self.proffession_number) == "87100":
            profession_name = "Дефектоскопист РГГ"
        elif str(self.proffession_number) == "87200":
            profession_name = "Дефектоскопист ПЗРС"
        elif str(self.proffession_number) == "08300":
            profession_name = "Фотолаборант"
        else:
            profession_name = "Неизвестный код"

        def PrintName(tabel):
            information = self.all_data["шифр"][str(self.proffession_number)]["Табельный"]
            name = F'{information[tabel]["фамилия"]} {information[tabel]["инициалы"]}'
            return name
        
        def PrintReason(reason):

            if reason == "ИО":
                print_reason = "И.о. мастера"
            elif reason == "О":
                print_reason = "Отпуск очередной"
            elif reason == "Э":
                print_reason = "Отпуск учебный"
            elif reason == "Р":
                print_reason = "Отпуск по беремености"
            elif reason == "А":
                print_reason = "Отпуск за свой счет"
            elif reason == "Ж":
                print_reason = "Пенсионный/уход за детьми"
            elif reason == "Д":
                print_reason = "Донорский день"
            elif reason == "М":
                print_reason = "Медкомиссия"
            elif reason == "Б":
                print_reason = "Больничный"
            elif reason == "К":
                print_reason = "Командировка"
            else:
                print_reason ="неизвестная причина"

            return print_reason
        
        def PrintPeriod(start_day, final_day):
            self.month = self.DATA.Current_month_int()
            self.year = self.DATA.Current_year_str()
            if int(start_day) != int(final_day):
                period = f"{int(start_day):02.0f}.{int(self.month):02.0f}-{int(final_day):02.0f}.{int(self.month):02.0f}.{int(self.year)}"
                
            elif int(start_day) == int(final_day):
                period = f"{int(final_day):02.0f}.{int(self.month):02.0f}.{int(self.year)}"
            
            else:
                period = f"Неизвестный период"
            
            return period
        
        def PrintCvalification(tabel): 
            cvalification = self.all_data["шифр"][str(self.proffession_number)]["Табельный"][tabel]["разряд"]
            return cvalification
        
        def CreateListForWriteXls():

            self.substitutes = self.DATA.Subtitutes()
            self.tarif = self.DATA.Tarif()
            list_for_write_xls = []

            for personal_number in self.tabels:
                substitutes_list = eval(self.substitutes["DEFAULT"][f"{self.proffession_number},{personal_number}"])

                for item in substitutes_list:
                    list_for_write_xls.append((
                        (f"{PrintName(item[0])} таб. {item[0]}"),
                        (f"{profession_name},{PrintCvalification(item[0])} разряд"),
                        (f"{PrintReason(item[1])}"),
                        (f"{PrintPeriod(item[2],item[3])}"),
                        (f"{PrintName(str(item[4]))}, таб {item[4]}"),
                        (f"{profession_name},{PrintCvalification(str(item[4]))} разряд"),
                        (f"{self.tarif[str(self.proffession_number)][PrintCvalification(item[0])]} ")
                                        ))

            return list_for_write_xls

        return CreateListForWriteXls()

class CREATE_EXCELL():
    def __init__(self, proffession_number,worksheet,workbook):
        self.proffession_number = str(proffession_number)
        self.workbook = workbook
        self.worksheet = worksheet
        super(CREATE_EXCELL, self).__init__()
        # подгружаем данные из функции- возвращает список замещения
        self.CDTWIE = CREATE_DATA_TO_WRITE_IN_EXCELL(self.proffession_number)
        self.DATA = LOAD_DATA(self.proffession_number)

    def Main(self):
        self.how_many_list_create()
        self.Create_document()
        self.Save_document()
        self.Xlsx_to_xls()
            
    def how_many_list_create(self):
        """Определяем сколько страниц в документе нужно создать"""
        # 40
        self.data_list = self.CDTWIE.Main()
        self.rows_count = len(self.data_list)
        # self.rows_count = 20
        self.pages_count = math.ceil(self.rows_count/11)
        return self.pages_count
    
    def Create_document(self):
        # создаем документ, с нужным количеством страниц
        self.start_row = 1
        for page in range(1,self.pages_count+1):
            CREATE_FORM(worksheet=self.worksheet,
                        start_row=self.start_row,
                        month = self.DATA.Current_month_str(),
                        year = self.DATA.Current_year_int(),
                        main_person=self.DATA.Current_main_person(),
                        main_person_name=self.DATA.Current_main_person_name(),
                        botiz=self.DATA.Current_botiz(),
                        botiz_name=self.DATA.Current_botiz_name(),
                        number_vedom=self.DATA.Current_vedomosti())
            self.start_row+=40
        
        self.current_row = 14
        count = 1
        
        # добавляем в этот документ нужное количество строк(обведенное границами)
        # заполняем строки
        for row in range(1,self.rows_count+1):
            self.write_border_row(current_row=self.current_row)
            self.write_text_row(current_row = self.current_row,count = count, value = self.data_list)
            count+=1
            if count == 12:
                self.current_row+=20
                count = 1
            else:
                self.current_row+=2
         
    def write_border_row(self,current_row):
        start_row = current_row
        self.worksheet.cell(start_row, 2).border = Border(left=Side(style='medium'),
													right=Side(style='medium'),
													top=Side(style='medium'),
													bottom=Side(style='medium'))
        self.worksheet.cell(start_row, 2).alignment = Alignment(horizontal='center',vertical='center')
        self.worksheet.cell(start_row, 2).font = Font(size=10, name='Times New Roman')
        
        self.worksheet.merge_cells(start_row = start_row,
						end_row = start_row+1,
						start_column = 2,
						end_column = 2)
		
		#фамилия
        self.worksheet.cell(start_row, 3).border = Border(left=Side(style='medium'),
                                                    right=Side(style='medium'),
                                                    top=Side(style='medium'),
                                                    bottom=Side(style='thin'))
        self.worksheet.cell(start_row, 3).font = Font(size=10, name='Times New Roman')
        self.worksheet.merge_cells(start_row = start_row,
                        end_row = start_row,
                        start_column = 3,
                        end_column = 4)

        #Профессия
        self.worksheet.cell(start_row+1, 3).border = Border(left=Side(style='medium'),
                                                    right=Side(style='medium'),
                                                    top=Side(style='thin'),
                                                    bottom=Side(style='medium'))
        self.worksheet.cell(start_row+1, 3).font = Font(size=10, name='Times New Roman')
        self.worksheet.merge_cells(start_row = start_row+1,
                        end_row = start_row+1,
                        start_column = 3,
                        end_column = 4)

        # Причина оотсутствия
        self.worksheet.cell(start_row, 5).border = Border(left=Side(style='medium'),
                                                    right=Side(style='medium'),
                                                    top=Side(style='medium'),
                                                    bottom=Side(style=None))
        self.worksheet.cell(start_row, 5).font = Font(size=10, name='Times New Roman')
        self.worksheet.cell(start_row, 5).alignment = Alignment(horizontal='center',vertical='center')
        self.worksheet.merge_cells(start_row = start_row,
                        end_row = start_row,
                        start_column = 5,
                        end_column = 6)
        # Период отстутсвия
        self.worksheet.cell(start_row+1, 5).border = Border(left=Side(style='medium'),
                                                    right=Side(style='medium'),
                                                    top=Side(style=None),
                                                    bottom=Side(style='medium'))
        self.worksheet.cell(start_row+1, 5).font = Font(size=10, name='Times New Roman')
        self.worksheet.cell(start_row+1, 5).alignment = Alignment(horizontal='center',vertical='center')
        self.worksheet.merge_cells(start_row = start_row+1,
                        end_row = start_row+1,
                        start_column = 5,
                        end_column = 6)
        #Тариф
        self.worksheet.cell(start_row, 7).border = Border(left=Side(style='medium'),
                                                    right=Side(style='medium'),
                                                    top=Side(style='medium'),
                                                    bottom=Side(style='medium'))
        self.worksheet.cell(start_row, 7).font = Font(size=10, name='Times New Roman')
        self.worksheet.cell(start_row, 7).alignment = Alignment(horizontal='center',vertical='center')
        self.worksheet.merge_cells(start_row = start_row,
                        end_row = start_row+1,
                        start_column = 7,
                        end_column = 7)
        #фамилия
        self.worksheet.cell(start_row, 8).border = Border(left=Side(style='medium'),
                                                    right=Side(style='medium'),
                                                    top=Side(style='medium'),
                                                    bottom=Side(style='thin'))
        self.worksheet.cell(start_row, 8).font = Font(size=10, name='Times New Roman')
        self.worksheet.merge_cells(start_row = start_row,
                        end_row = start_row,
                        start_column = 8,
                        end_column = 10)

        #Профессия
        self.worksheet.cell(start_row+1, 8).border = Border(left=Side(style='medium'),
                                                    right=Side(style='medium'),
                                                    top=Side(style='thin'),
                                                    bottom=Side(style='medium'))
        self.worksheet.cell(start_row+1, 8).font = Font(size=10, name='Times New Roman')
        self.worksheet.merge_cells(start_row = start_row+1,
                        end_row = start_row+1,
                        start_column = 8,
                        end_column = 10)
        # Процент
        self.worksheet.cell(start_row, 11).border = Border(left=Side(style='medium'),
                                                    right=Side(style='medium'),
                                                    top=Side(style='medium'),
                                                    bottom=Side(style='medium'))
        self.worksheet.cell(start_row, 11).font = Font(size=10, name='Times New Roman')
        self.worksheet.cell(start_row, 11).alignment = Alignment(horizontal='center',vertical='center')
        self.worksheet.merge_cells(start_row = start_row,
                        end_row = start_row+1,
                        start_column = 11,
                        end_column = 11)
        # согласие
        self.worksheet.cell(start_row, 12).border = Border(left=Side(style='medium'),
                                                    right=Side(style='medium'),
                                                    top=Side(style='medium'),
                                                    bottom=Side(style='medium'))
        self.worksheet.merge_cells(start_row = start_row,
                        end_row = start_row+1,
                        start_column = 12,
                        end_column = 12)
        # подпись
        self.worksheet.cell(start_row, 13).border = Border(left=Side(style='medium'),
                                                    right=Side(style='medium'),
                                                    top=Side(style='medium'),
                                                    bottom=Side(style='medium'))
        self.worksheet.merge_cells(start_row = start_row,
                        end_row = start_row+1,
                        start_column = 13,
                        end_column = 13)

    def write_text_row(self,current_row,count,value):

        self.worksheet.cell(current_row, 2,str(count))
        self.worksheet.cell(current_row, 3,value[count-1][0])
        self.worksheet.cell(current_row+1, 3,value[count-1][1])
        self.worksheet.cell(current_row, 5,value[count-1][2])
        self.worksheet.cell(current_row+1, 5,value[count-1][3])
        self.worksheet.cell(current_row, 7,value[count-1][-1])
        self.worksheet.cell(current_row, 8,value[count-1][4])
        self.worksheet.cell(current_row+1, 8,value[count-1][5])
        self.worksheet.cell(current_row, 11,f"{self.DATA.Procent()[str(self.proffession_number)]}%")
        
    def Save_document(self):
        current_place = self.DATA.Current_place()
        current_year = self.DATA.Current_year_int()
        current_month = self.DATA.Current_month_str()
        current_month_int = self.DATA.Current_month_int()

        if not os.path.exists(f'ведомости\\{current_place}\\{current_year}\\{current_month}'):
                os.makedirs(os.path.join(("ведомости"),(f"{current_place}"),(f"{current_year}"),(f"{current_month}")))
        self.path_xlsx = f'ведомости\\{current_place}\\{current_year}\\{current_month}\\{self.proffession_number}_199_{current_year}_{current_month_int}.xlsx'
        self.workbook.save(self.path_xlsx)

    def Xlsx_to_xls(self):
        xlApp = Dispatch('Excel.Application')
        path = os.path.abspath (self.path_xlsx)
        wb = xlApp.Workbooks.open(path)
        wb.SaveAs(path[:-1], FileFormat=56)
        xlApp.Quit()
        os.remove(f"{path}")
if __name__ == "__main__":
    # CDTWIE = CREATE_DATA_TO_WRITE_IN_EXCELL("87100")
    # x = CDTWIE.Main()
    # print(x)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    CE = CREATE_EXCELL(proffession_number="87100",worksheet=worksheet,workbook=workbook)
    CE.Main()
    

    




