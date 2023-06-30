import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

class CREATE_FORM():
	def __init__(self,worksheet,
                start_row = 1,
                month = "январь",
                year = 0000,
                number_vedom=0,
                main_person = 'Начальник НИТИЦ',
                main_person_name = 'А.И. Власов',
                botiz = 'Начальник БОТиЗ',
                botiz_name = 'Н.А.Львова'):
		super(CREATE_FORM,self).__init__()
		self.worksheet = worksheet
		self.start_row = start_row
		self.month = month
		self.year = year
		self.number_vedom = number_vedom
		self.main_person = main_person
		self.main_person_name = main_person_name
		self.botiz = botiz
		self.botiz_name = botiz_name
		self.MAIN()
    
	def MAIN(self):
		self.create_top()
		self.create_work_table()
		self.create_bottom()
		self.width_column()
	
	def create_top(self):

		list_month_with_sfx = {'январь':'января',
                      'февраль' : 'февраля',
                      'март' : 'марта',
                      'апрель' :'апреля',
                      'май' : 'мая',
                      'июнь' : 'июня',
                      'июль' : 'июля',
                      'август' : 'августа',
                      'сентябрь' : 'сентября',
                      'октябрь' : 'октября',
                      'ноябрь' : 'ноября',
                      'декабрь' : 'декабря'}
		for key, value in list_month_with_sfx.items():
			if key == self.month:
				month_with_sfx = value

		_Font = Font(size = 10,name = 'Times New Roman')
		_Alignment =Alignment(wrap_text= True,horizontal='center',vertical='center')
		_Medium_border = Border(left=Side(style='medium'),
                     right=Side(style='medium'),
                     top=Side(style='medium'),
                     bottom=Side(style='medium'))
		
		
		value_1 = 'АО  "ПО "СЕВМАШ"'
		self.worksheet.cell(self.start_row, 2, value_1).font = _Font

		value_2 = 'Цех (отдел)      НИТИЦ'
		self.worksheet.cell(self.start_row+2, 2, value_2).font = _Font

		value_3 = 'СОГЛАШЕНИЕ-РАСЧЕТ ОПЛАТЫ' 
		self.worksheet.cell(self.start_row+4, 2, value_3).font = _Font

		value_4 = f'от  01 {month_with_sfx} {self.year} г. № 25.17/{self.number_vedom}    за {self.month} {self.year} г.'
		self.worksheet.cell(self.start_row+6, 2, value_4).font = _Font

		value_5 = '"ЗА ИСПОЛНЕНИЕ ОБЯЗАННОСТЕЙ ВРЕМЕННО ОТСУТСТВУЮЩЕГО РАБОТНИКА"'
		self.worksheet.cell(self.start_row, 5, value_5).font = _Font

		value_6 = "Руководствуясь положение 56.61-1.01.014.-2020, возложить частичное исполнение обязанностей временно"
		self.worksheet.cell(self.start_row+1, 5, value_6).font = _Font

		value_7 = 'отсутствующего работника с установлением оплаты видом 199 ведомостью РВО на следующих работников:'
		self.worksheet.cell(self.start_row+2, 5, value_7).font = _Font

		value_8 = '№ п/п'
		#sheet.merge(top_row, bottom_row, left_column, right_column)
		self.worksheet.cell(self.start_row+8, 2, value_8).font = _Font
		self.worksheet.cell(self.start_row+8, 2).alignment = _Alignment
		self.worksheet.cell(self.start_row+8, 2).border = _Medium_border
		self.worksheet.merge_cells(start_row = self.start_row+8,
							  end_row = self.start_row+8+4,
							  start_column = 2,
							  end_column = 2)
		
		value_9 = 'Отсутствующий работник, вакантная должность'
		self.worksheet.cell(self.start_row+8, 3, value_9).font = _Font
		self.worksheet.cell(self.start_row+8, 3).alignment = _Alignment
		self.worksheet.cell(self.start_row+8, 3).border = _Medium_border
		self.worksheet.merge_cells(start_row = self.start_row+8,
							  end_row = self.start_row+8,
							  start_column = 3,
							  end_column = 7)
		

		value_10 = 'Работник исполняющий обязанности временно отсутствующего'
		self.worksheet.cell(self.start_row+8, 8, value_10).font = _Font
		self.worksheet.cell(self.start_row+8, 8).alignment = _Alignment
		self.worksheet.cell(self.start_row+8, 8).border = _Medium_border
		self.worksheet.merge_cells(start_row = self.start_row+8,
							  end_row = self.start_row+8,
							  start_column = 8,
							  end_column = 13)

		value_11 = 'Фамилия И.О., таб. №	'
		self.worksheet.cell(self.start_row+9, 3, value_11).font = _Font
		self.worksheet.cell(self.start_row+9, 3).alignment = _Alignment
		self.worksheet.cell(self.start_row+9, 3).border = _Medium_border
		self.worksheet.merge_cells(start_row = self.start_row+9,
							  end_row = self.start_row+10,
							  start_column = 3,
							  end_column = 4)

		value_12 = 'Профессия (должность),  разряд'
		self.worksheet.cell(self.start_row+11, 3, value_12).font = _Font
		self.worksheet.cell(self.start_row+11, 3).alignment = _Alignment
		self.worksheet.cell(self.start_row+11, 3).border = _Medium_border
		self.worksheet.merge_cells(start_row = self.start_row+11,
							  end_row = self.start_row+12,
							  start_column = 3,
							  end_column = 4)

		value_13 = 'Причина отсутствия'
		self.worksheet.cell(self.start_row+9, 5, value_13).font = _Font
		self.worksheet.cell(self.start_row+9, 5).alignment = _Alignment
		self.worksheet.cell(self.start_row+9, 5).border = _Medium_border
		self.worksheet.merge_cells(start_row = self.start_row+9,
							  end_row = self.start_row+10,
							  start_column = 5,
							  end_column = 6)

		value_14 = 'Период отсутствия'
		self.worksheet.cell(self.start_row+11, 5, value_14).font = _Font
		self.worksheet.cell(self.start_row+11, 5).alignment = _Alignment
		self.worksheet.cell(self.start_row+11, 5).border = _Medium_border
		self.worksheet.merge_cells(start_row = self.start_row+11,
							  end_row = self.start_row+12,
							  start_column = 5,
							  end_column = 6)

		value_15 = 'Тариф (оклад), руб'
		self.worksheet.cell(self.start_row+9, 7, value_15).font = _Font
		self.worksheet.cell(self.start_row+9, 7).alignment = _Alignment
		self.worksheet.cell(self.start_row+9, 7).border = _Medium_border
		self.worksheet.merge_cells(start_row = self.start_row+9,
							  end_row = self.start_row+12,
							  start_column = 7,
							  end_column = 7)

		value_16 = 'Фамилия И.О., таб. №'
		self.worksheet.cell(self.start_row+9, 8, value_16).font = _Font
		self.worksheet.cell(self.start_row+9, 8).alignment = _Alignment
		self.worksheet.cell(self.start_row+9, 8).border = _Medium_border
		self.worksheet.merge_cells(start_row = self.start_row+9,
							  end_row = self.start_row+10,
							  start_column = 8,
							  end_column = 10)

		value_17 = 'Профессия (должность), разряд'
		self.worksheet.cell(self.start_row+11, 8, value_17).font = _Font
		self.worksheet.cell(self.start_row+11, 8).alignment = _Alignment
		self.worksheet.cell(self.start_row+11, 8).border = _Medium_border
		self.worksheet.merge_cells(start_row = self.start_row+11,
								end_row = self.start_row+12,
								start_column = 8,
								end_column = 10)

		value_18 = 'Предварительный % оплаты от тарифа (оклада) отсутствующего'
		self.worksheet.cell(self.start_row+9, 11, value_18).font = _Font
		self.worksheet.cell(self.start_row+9, 11).alignment = _Alignment
		self.worksheet.cell(self.start_row+9, 11).border = _Medium_border
		self.worksheet.merge_cells(start_row = self.start_row+9,
								end_row = self.start_row+12,
								start_column = 11,
								end_column = 11)

		value_19 = 'Согласие на исполнение, подпись, дата'
		self.worksheet.cell(self.start_row+9, 12, value_19).font = _Font
		self.worksheet.cell(self.start_row+9, 12).alignment = _Alignment
		self.worksheet.cell(self.start_row+9, 12).border = _Medium_border
		self.worksheet.merge_cells(start_row = self.start_row+9,
								end_row = self.start_row+12,
								start_column = 12,
								end_column = 12)

		value_20 = 'Окончательный размер оплаты, руб.'
		self.worksheet.cell(self.start_row+9, 13, value_20).font = _Font
		self.worksheet.cell(self.start_row+9, 13).alignment = _Alignment
		self.worksheet.cell(self.start_row+9, 13).border = _Medium_border
		self.worksheet.merge_cells(start_row = self.start_row+9,
								end_row = self.start_row+12,
								start_column = 13,
								end_column = 13)
	
	def create_row(self,start_row):
		# номер п/п
		self.worksheet.cell(start_row, 2).border = Border(left=Side(style='medium'),
													right=Side(style='medium'),
													top=Side(style='medium'),
													bottom=Side(style='medium'))
		self.worksheet.merge_cells(start_row = start_row,
						end_row = start_row+1,
						start_column = 2,
						end_column = 2)
		
		#фамилия
		self.worksheet.cell(start_row, 3).border = Border(left=Side(style='medium'),
													right=Side(style='medium'),
													top=Side(style='medium'),
													bottom=Side(style='thin'))
		self.worksheet.merge_cells(start_row = start_row,
						end_row = start_row,
						start_column = 3,
						end_column = 4)
		
		#Профессия
		self.worksheet.cell(start_row+1, 3).border = Border(left=Side(style='medium'),
													right=Side(style='medium'),
													top=Side(style='thin'),
													bottom=Side(style='medium'))
		self.worksheet.merge_cells(start_row = start_row+1,
						end_row = start_row+1,
						start_column = 3,
						end_column = 4)
		
		# Причина оотсутствия
		self.worksheet.cell(start_row, 5).border = Border(left=Side(style='medium'),
													right=Side(style='medium'),
													top=Side(style='medium'),
													bottom=Side(style=None))
		self.worksheet.merge_cells(start_row = start_row,
						end_row = start_row,
						start_column = 5,
						end_column = 6)
		# Период отстутсвия
		self.worksheet.cell(start_row+1, 5).border = Border(left=Side(style='medium'),
													right=Side(style='medium'),
													top=Side(style=None),
													bottom=Side(style='medium'))
		self.worksheet.merge_cells(start_row = start_row+1,
						end_row = start_row+1,
						start_column = 5,
						end_column = 6)
		#Тариф
		self.worksheet.cell(start_row, 7).border = Border(left=Side(style='medium'),
													right=Side(style='medium'),
													top=Side(style='medium'),
													bottom=Side(style='medium'))
		self.worksheet.merge_cells(start_row = start_row,
						end_row = start_row+1,
						start_column = 7,
						end_column = 7)
		#фамилия
		self.worksheet.cell(start_row, 8).border = Border(left=Side(style='medium'),
													right=Side(style='medium'),
													top=Side(style='medium'),
													bottom=Side(style='thin'))
		self.worksheet.merge_cells(start_row = start_row,
						end_row = start_row,
						start_column = 8,
						end_column = 10)
		
		#Профессия
		self.worksheet.cell(start_row+1, 8).border = Border(left=Side(style='medium'),
													right=Side(style='medium'),
													top=Side(style='thin'),
													bottom=Side(style='medium'))
		self.worksheet.merge_cells(start_row = start_row+1,
						end_row = start_row+1,
						start_column = 8,
						end_column = 10)
		# Процент
		self.worksheet.cell(start_row, 11).border = Border(left=Side(style='medium'),
													right=Side(style='medium'),
													top=Side(style='medium'),
													bottom=Side(style='medium'))
		self.worksheet.merge_cells(start_row = start_row,
						end_row = start_row+1,
						start_column = 11,
						end_column = 11)
		# согласие
		self.worksheet.cell(start_row, 12).border = Border(left=Side(style='medium'),
													right=Side(style='medium'),
													top=Side(style='medium'),
													bottom=Side(style='medium'))
		self.worksheet.merge_cells(start_row = start_row,
						end_row = start_row+1,
						start_column = 12,
						end_column = 12)
		# подпись
		self.worksheet.cell(start_row, 13).border = Border(left=Side(style='medium'),
													right=Side(style='medium'),
													top=Side(style='medium'),
													bottom=Side(style='medium'))
		self.worksheet.merge_cells(start_row = start_row,
						end_row = start_row+1,
						start_column = 13,
						end_column = 13)

	def create_work_table(self):
		self.cv = self.create_row(start_row = 14)
	
	def create_bottom(self):

		x_value_21 = len(f'{self.main_person}/{self.main_person_name} / Мастер (руководитель работ)/{self.botiz}/{self.botiz_name}/')
		#Собираем строку
		
		value_21 =f'{self.main_person}{int((142-x_value_21)/4)*"_"}/{self.main_person_name} / Мастер (руководитель работ){int((135-x_value_21)/4)*"_"}/{int((135-x_value_21)/4)*"_"}{self.botiz}{int((135-x_value_21)/4)*"_"}/{self.botiz_name}/'
		self.worksheet.cell(self.start_row+37, 2, value_21).alignment = Alignment(horizontal='left',vertical='center')
		self.worksheet.cell(self.start_row+37, 2, value_21).font = Font(size = 10,name = 'Times New Roman')
		
		value_22 =f'{34*" "}подпись,дата{80*" "}подпись,дата{56*" "}подпись,дата'
		self.worksheet.cell(self.start_row+38, 2, value_22).alignment = Alignment(horizontal='left',vertical='center')
		self.worksheet.cell(self.start_row+38, 2, value_22).font = Font(size = 10,name = 'Times New Roman')
		
		value_23 = 'ф.56.61.70'
		self.worksheet.cell(self.start_row+39, 13, value_23).alignment = Alignment(horizontal='right',vertical='center')
		self.worksheet.cell(self.start_row+39, 13, value_23).font = Font(size = 10,name = 'Times New Roman')
					
	def width_column(self):
		self.worksheet.column_dimensions[get_column_letter(1)].width = 3.14+0.71
		self.worksheet.column_dimensions[get_column_letter(2)].width = 4+0.71
		self.worksheet.column_dimensions[get_column_letter(3)].width = 10.57+0.71
		self.worksheet.column_dimensions[get_column_letter(4)].width = 16.57+0.71
		self.worksheet.column_dimensions[get_column_letter(5)].width = 9.71+0.71
		self.worksheet.column_dimensions[get_column_letter(6)].width = 6.86+0.71
		self.worksheet.column_dimensions[get_column_letter(7)].width = 7.29+0.71
		self.worksheet.column_dimensions[get_column_letter(8)].width = 8.43+0.71
		self.worksheet.column_dimensions[get_column_letter(9)].width = 8.43+0.71
		self.worksheet.column_dimensions[get_column_letter(9)].width = 9.71+0.71

				


		


				
		

    
if __name__ == '__main__':
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    CF = CREATE_FORM(worksheet=worksheet)
    workbook.save("test.xlsx")
        