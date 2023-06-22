import sys
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

class CREATE_FORM():
    def __init__(self,worksheet, start_row,start_column, 
                brigadir,
                left_day, right_day,
                month, year):
        super(CREATE_FORM,self).__init__()

        self.worksheet = worksheet
        self.start_row = start_row
        self.start_column = start_column
        self.brigadir = brigadir
        self.left_day = left_day
        self.right_day = right_day
        self.month = month
        self.year = year
        
        self.MAIN()

    def MAIN(self):
        # левая верхняяя часть страницы
        self.create_top_text(start_row=self.start_row,
                             start_column=self.start_column,
                             day=self.left_day)
        # правая верхняя часть страницы 
        self.create_top_text(start_row=self.start_row,
                             start_column=self.start_column+8,
                             day = self.right_day)
        self.create_table(start_row=self.start_row,
                          start_column=self.start_column)
        
		# таблица слева снизу
        self.create_table(start_row=self.start_row+18,
                          start_column=self.start_column)
        
        # таблица справа сверху
        self.create_table(start_row=self.start_row,
                          start_column=self.start_column+8)
        
        # таблица справа снизу
        self.create_table(start_row=self.start_row+18,
                          start_column=self.start_column+8)
        
        # текст внизу левой таблицы
        self.create_bottom_text(start_row=self.start_row-1,
                                start_column=self.start_column)
        
        # текст внизу правой таблицы
        self.create_bottom_text(start_row=self.start_row-1,
                                start_column=self.start_column+8)
        self.number_form(start_row=self.start_row,
                         start_column=self.start_column)
        self.number_form(start_row=self.start_row,
                         start_column=self.start_column+8)
        self.number_form(start_row=self.start_row+25,start_column=self.start_column)

        self.number_form(start_row=self.start_row+25,start_column=self.start_column+8)

        # устанавливаем ширину колонок
        self.width_column(start_column=self.start_column) #колонки левой части
        self.width_column(start_column=self.start_column+8)

        # настройки листа
        self.create_settings_list()
        
    def create_top_text(self,start_row,start_column,day):

        value_1 = "АО «ПО «Севмаш»"
        self.worksheet.cell(start_row,start_column, value = value_1).font = Font(size = 11,
                                                                                name = 'Times New Roman')
        self.worksheet.merge_cells(start_row = start_row,
                                   start_column = start_column,
                                   end_row = start_row, 
                                   end_column = start_column+1)
        

        value_2 = "УТВЕРЖДАЮ"
        self.worksheet.cell(start_row,start_column+3,value = value_2).font = Font(size = 11,
                                                                                  bold = True,
                                                                                  name = 'Times New Roman')
        self.worksheet.cell(start_row,start_column+3).alignment = Alignment(horizontal='center')
        self.worksheet.merge_cells(start_row = start_row,
                                   start_column =start_column+3,
                                   end_row=start_row,
                                   end_column=start_column+5)
        

        value_3 = "  Руководитель подразделения"
        self.worksheet.cell(start_row+1,start_column+3,value = value_3).font = Font(size = 11,
                                                                                     bold= True,
                                                                                     name='Times New Roman')
        self.worksheet.cell(start_row+1,start_column+3).alignment = Alignment(horizontal='center')
        self.worksheet.merge_cells(start_row = start_row+1,
                                   end_row = start_row+1,
                                   start_column = start_column+3,
                                   end_column = start_column+5)
        self.worksheet.row_dimensions[start_row+1].height = 13


        value_4 = "_____________________________________________________"
        self.worksheet.cell(start_row+2,start_column+3,value = value_4).font = Font(size = 7,
                                                                                    name='Times New Roman')
        self.worksheet.merge_cells(start_row = start_row+2,
                                   end_row = start_row+2,
                                   start_column = start_column+3,
                                   end_column = start_column+5)
        self.worksheet.row_dimensions[start_row+2].height = 13


        value_5 = "(должность)"
        self.worksheet.cell(start_row+3,start_column+3,value_5).font = Font(size = 8,
                                                                            name='Times New Roman')
        self.worksheet.cell(start_row+3,start_column+3,value_5).alignment = Alignment(horizontal='center', 
                                                                                      vertical='center')
        self.worksheet.merge_cells(start_row = start_row+3,
                                   end_row = start_row+3,
                                   start_column = start_column+3,
                                   end_column = start_column+5)
        self.worksheet.row_dimensions[start_row+3].height = 8
    

        value_6 = "________________   __________________________________"
        self.worksheet.cell(start_row+4,start_column+3,value_6).font = Font(size = 7,
                                                                            name='Times New Roman')
        self.worksheet.cell(start_row+4,start_column+3,value_6).alignment = Alignment(horizontal='center', 
                                                                                      vertical='center')
        self.worksheet.merge_cells(start_row = start_row+4,
                                   end_row = start_row+4,
                                   start_column = start_column+3,
                                   end_column = start_column+5)
        self.worksheet.row_dimensions[start_row+4].height = 13
        

        value_7 = "(подпись, дата)     (Расшифровка подписи)"
        self.worksheet.cell(start_row+5,start_column+3,value_7).font = Font(size = 8,
                                                                            name='Times New Roman')
        self.worksheet.cell(start_row+5,start_column+3,value_7).alignment = Alignment(horizontal='left',
                                                                                      vertical='center')
        self.worksheet.merge_cells(start_row = start_row+5,
                                   end_row = start_row+5,
                                   start_column = start_column+3,
                                   end_column = start_column+5)
        self.worksheet.row_dimensions[start_row+5].height = 9
        

        value_8 = "ДЛЯ 5-ДНЕВНОЙ РАБОЧЕЙ НЕДЕЛИ"
        self.worksheet.cell(start_row+6,start_column,value_8).font = Font(size = 9,
                                                                          bold = True,
                                                                          name='Times New Roman')
        self.worksheet.cell(start_row+6,start_column,value_8).alignment = Alignment(horizontal='center')
        self.worksheet.merge_cells(start_row = start_row+6,
                                   end_row = start_row+6,
                                   start_column = start_column,
                                   end_column = start_column+5)
        self.worksheet.row_dimensions[start_row+6].height = 15

        value_9 = f'Цех №25     Бригада {self.brigadir}     заказ 981760'
        self.worksheet.cell(start_row+7,start_column,value_9).font = Font(size = 12,
                                                                          name='Times New Roman')
        self.worksheet.cell(start_row+7,start_column,value_9).alignment = Alignment(horizontal='center')
        self.worksheet.merge_cells(start_row = start_row+7,
                                   end_row = start_row+7,
                                   start_column = start_column,
                                   end_column = start_column+5)
        self.worksheet.row_dimensions[start_row+7].height = 15

        value_10 = 'СПИСОК'
        self.worksheet.cell(start_row+8,start_column,value_10).font = Font(size = 14,
                                                                          bold = True,
                                                                          name='Times New Roman')
        self.worksheet.cell(start_row+8,start_column,value_10).alignment = Alignment(horizontal='center',vertical='center')
        self.worksheet.merge_cells(start_row = start_row+8,
                                   end_row = start_row+8,
                                   start_column = start_column,
                                   end_column = start_column+5)
        self.worksheet.row_dimensions[start_row+8].height = 15
        
        
        value_11 = f"на выдачу талонов на «{day}» {self.month}  {self.year} г."
        self.worksheet.cell(start_row+9,start_column,value_11).font = Font(size = 11,
                                                                          bold = True,
                                                                          name='Times New Roman')
        self.worksheet.cell(start_row+9,start_column,value_11).alignment = Alignment(horizontal='center',vertical='center')
        self.worksheet.merge_cells(start_row = start_row+9,
                                   end_row = start_row+9,
                                   start_column = start_column,
                                   end_column = start_column+5)
        self.worksheet.row_dimensions[start_row+9].height = 18
        
    def create_table(self,start_row,start_column):

        Medium_border = Border(left=Side(style='medium'),
                     right=Side(style='medium'),
                     top=Side(style='medium'),
                     bottom=Side(style='medium'))
        
        _Font = Font(size = 8,name = 'Times New Roman')
        _Alignment =Alignment(wrap_text= True,horizontal='center',vertical='center')
            
        value_12 = "№ п/п"
        self.worksheet.cell(start_row+10,start_column,value_12).font = _Font
        self.worksheet.cell(start_row+10,start_column,value_12).alignment = _Alignment
        self.worksheet.cell(start_row+10,start_column,value_12).border = Border(left=Side(style=None),
                                                                                right=Side(style='medium'),
                                                                                top=Side(style='medium'),
                                                                                bottom=Side(style='medium'))
    
        value_13 = 'Фамилия, Имя, Отчество'
        self.worksheet.cell(start_row+10,start_column+1,value_13).font = _Font
        self.worksheet.cell(start_row+10,start_column+1,value_13).alignment = _Alignment
        self.worksheet.cell(start_row+10,start_column+1,value_13).border = Medium_border
        

        value_14 = 'Профессия, должность'
        self.worksheet.cell(start_row+10,start_column+2,value_14).font = _Font
        self.worksheet.cell(start_row+10,start_column+2,value_14).alignment = _Alignment
        self.worksheet.cell(start_row+10,start_column+2,value_14).border = Medium_border
        
        value_15 = 'табельный номер'
        self.worksheet.cell(start_row+10,start_column+3,value_15).font = _Font
        self.worksheet.cell(start_row+10,start_column+3,value_15).alignment = _Alignment
        self.worksheet.cell(start_row+10,start_column+3,value_15).border = Medium_border
        
        value_16 = 'кол-во талонов'
        self.worksheet.cell(start_row+10,start_column+4,value_16).font = _Font
        self.worksheet.cell(start_row+10,start_column+4,value_16).alignment = _Alignment
        self.worksheet.cell(start_row+10,start_column+4,value_16).border = Medium_border
        
        value_17 = 'Подпись в получении (лично)'
        self.worksheet.cell(start_row+10,start_column+5,value_17).font = _Font
        self.worksheet.cell(start_row+10,start_column+5,value_17).alignment = _Alignment
        self.worksheet.cell(start_row+10,start_column+5,value_17).border = Border(left=Side(style='medium'),
                                                                                  right=Side(style=None),
                                                                                  top=Side(style='medium'),
                                                                                  bottom=Side(style='medium'))
        self.worksheet.row_dimensions[start_row+10].height = 35


        # крайние столбцы таблицы
        for row in range(11,27): 
            self.worksheet.cell(start_row+row,start_column).border = Border(bottom=Side(style='thin'))
            self.worksheet.cell(start_row+row,start_column+5).border = Border(bottom=Side(style='thin'))
        # границы таблицы
        for row in range(11,27):
            for column in range(1,5):
                self.worksheet.cell(start_row+row,start_column+column).border = Border(left=Side(style='thin'),
                                                                                  right=Side(style='thin'),
                                                                                  bottom=Side(style='thin'))
        # устанавливаем высоту строк
        for i in range(11,27):
            self.worksheet.row_dimensions[start_row+i].height = 23
        
    def create_bottom_text(self,start_row,start_column):

        value_21 = "    Должностные лица несут полную материальную ответсвенность за выдачу ими талонов на питание работникам, не выполняющим работ в районах и на работах, где по заключению отдела ОЯРБ должно предоставляться бесплатное питание."
        self.worksheet.merge_cells(start_row = start_row+45,
                                   end_row = start_row+45,
                                   start_column = start_column,
                                   end_column = start_column+5)
        self.worksheet.cell(start_row+45,start_column,value_21).font = Font(size = 11, name = 'Times New Roman')
        self.worksheet.cell(start_row+45,start_column).alignment = Alignment(wrap_text=True, horizontal="justify")
        
        for column in range(0,6):
            self.worksheet.cell(start_row+45,start_column+column).border = Border(bottom=Side(style=None),
                                                                                  top = Side(style = None))
        
        self.worksheet.row_dimensions[start_row+45].height = 63

        value_22 = "Всего получено ___________________________________________________________"
        self.worksheet.merge_cells(start_row = start_row+46,
                                   end_row = start_row+46,
                                   start_column = start_column,
                                   end_column = start_column+5)
        self.worksheet.cell(start_row+46,start_column,value_22).font = Font(size = 11, name = 'Times New Roman')
        self.worksheet.row_dimensions[start_row+46].height = 18
        
        value_23 = "Выдано ___________________________________________________________________"
        self.worksheet.merge_cells(start_row = start_row+47,
                                   end_row = start_row+47,
                                   start_column = start_column,
                                   end_column = start_column+5)
        self.worksheet.cell(start_row+47,start_column,value_23).font = Font(size = 11, name = 'Times New Roman')
        self.worksheet.row_dimensions[start_row+47].height = 18
        
        value_24 = "Сдан  остаток _____________________________________________________________"
        self.worksheet.merge_cells(start_row = start_row+48,
                                   end_row = start_row+48,
                                   start_column = start_column,
                                   end_column = start_column+5)
        self.worksheet.cell(start_row+48,start_column,value_24).font = Font(size = 11, name = 'Times New Roman')
        self.worksheet.row_dimensions[start_row+48].height = 18

        value_25 = "Выдал _____________________________________________________________________"
        self.worksheet.merge_cells(start_row = start_row+49,
                                   end_row = start_row+49,
                                   start_column = start_column,
                                   end_column = start_column+5)
        self.worksheet.cell(start_row+49,start_column,value_25).font = Font(size = 11, name = 'Times New Roman')
        self.worksheet.row_dimensions[start_row+49].height = 18

        value_26 = "(должность)                                (подпись)                    (расшифровка подписи)"
        self.worksheet.merge_cells(start_row = start_row+50,
                                   end_row = start_row+50,
                                   start_column = start_column,
                                   end_column = start_column+5)
        self.worksheet.cell(start_row+50,start_column,value_26).font = Font(size = 8, name = 'Times New Roman')
        self.worksheet.cell(start_row+50,start_column,value_26).alignment = Alignment(horizontal='right',vertical='center')
        self.worksheet.row_dimensions[start_row+50].height = 8
    
        value_27 = 'Проверил факт работы дозиметрист ___________________________________________'
        self.worksheet.merge_cells(start_row = start_row+51,
                                   end_row = start_row+51,
                                   start_column = start_column,
                                   end_column= start_column+5)
        self.worksheet.cell(start_row+51,start_column,value_27).font = Font(size = 11, name = 'Times New Roman')
        self.worksheet.row_dimensions[start_row+51].height = 18

        value_28 = "(подпись)                    (расшифровка подписи)"
        self.worksheet.merge_cells(start_row= start_row+52,
                                   end_row= start_row+52,
                                   start_column = start_column,
                                   end_column = start_column+5)
        self.worksheet.cell(start_row+52,start_column,value_28).font = Font(size = 8, name = 'Times New Roman')
        self.worksheet.cell(start_row+52,start_column,value_28).alignment = Alignment(horizontal='right',vertical='center')
        self.worksheet.row_dimensions[start_row+52].height = 8

    def number_form(self,start_column,start_row):
        value_20 = "Ф.27.62.02"	
        self.worksheet.cell(start_row+27,start_column+5,value_20).font = Font(size = 10, name = "Times New Roman")

    def width_column(self,start_column):

        self.worksheet.column_dimensions[get_column_letter(start_column)].width = 4+0.71
        self.worksheet.column_dimensions[get_column_letter(start_column+1)].width = 17+0.71
        self.worksheet.column_dimensions[get_column_letter(start_column+2)].width = 11+0.71
        self.worksheet.column_dimensions[get_column_letter(start_column+3)].width = 9+0.71
        self.worksheet.column_dimensions[get_column_letter(start_column+4)].width = 9+0.71
        self.worksheet.column_dimensions[get_column_letter(start_column+5)].width = 9+0.71
        self.worksheet.column_dimensions[get_column_letter(start_column+6)].width = 6+0.71
        self.worksheet.column_dimensions[get_column_letter(start_column+7)].width = 6+0.71

    def create_settings_list(self):
        self.worksheet.page_setup.orientation = 'landscape'
        self.worksheet.page_setup.paperSize = self.worksheet.PAPERSIZE_A4
        self.worksheet.page_margins = PageMargins(left=0, right=0, top=0.3, bottom=0.3, header=20, footer=20)
        self.worksheet.print_options.horizontalCentered = True
        self.worksheet.print_options.verticalCentered = True
      
if __name__ == "__main__":
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        CF = CREATE_FORM(worksheet=worksheet,start_row = 1,start_column = 1, 
                        brigadir = "Илюхин Ю.В", 
                        left_day = "1",right_day = "__",
                        month = "________", year = "__")
        workbook.save("Talons_2\For_exel\\form.xlsx")
        