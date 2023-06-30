import os
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from win32com.client import Dispatch
from Create_form import CREATE_FORM
from Function_for_create_talons_list import ALL_FUNCTION_TO_CREATE_TALONS_LIST as AFTCTL


class PRINT_TO_XLS:
    def __init__(self,start_data,final_data,month = "месяц",year = "год",brigadir = "бригадир",place = "неизвестный цех"):
        super(PRINT_TO_XLS,self).__init__()
        self.start_data = start_data
        self.final_data = final_data
        self.month = month
        self.year = year
        self.brigadir = brigadir
        self.place = place
        self.CREATE_FORM = CREATE_FORM
    def main(self):
        self.create_xls_file()
        self.write_to_xls()
        self.save_workbook()
    def create_xls_file(self):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        
    def write_to_xls(self):
        # подгружаем функцию, которая использует функцию из подгружаемого файла
        self.print_list = AFTCTL(start_data=self.start_data,final_data=self.final_data).create_list_with_lists_per_data()
        how_many_days = len(self.print_list)
        step = 53 # шаг с которым добавляются новые листы в последующую строчку после предыдущего
        start_row = 1
        # записываем данные в файл
        for days in range(0,int(how_many_days),2):
            try:
                # созданм форму используя подгружаемую функцию
                self.CREATE_FORM(worksheet=self.worksheet,start_row = start_row,start_column = 1, 
                            brigadir = self.brigadir, 
                            left_day = self.print_list[days][0][3],right_day = self.print_list[days+1][0][3],
                            month = self.Month_function(self.month), year = self.year)
                # заполняем форму использую функцию писатель
                self.writer(start_row,left=True,right=True,day=days)
                start_row+=step
            except IndexError:
                self.CREATE_FORM(worksheet=self.worksheet,start_row = start_row,start_column = 1, 
                            brigadir = self.brigadir, 
                            left_day = self.print_list[days][0][3],right_day = "__",
                            month = self.Month_function(self.month), year = self.year)
                self.writer(start_row,left=True,right=False,day=days)

    def Month_function(self,month):
        month = int(month)
        months = {1:"Январь",
                  2:"Февраль",
                  3: "Март",
                  4: "Апрель",
                  5: "Май",
                  6: "Июнь",
                  7: "Июль",
                  8: "Август",
                  9: "Сентябрь",
                  10: "Октябрь",
                  11: "Ноябрь",
                  12: "Декабрь"}
        return months[month]                
           
    def writer(self,start_row=1,left = False,right = False,day = 0):
        # функция писаттель, которая записывает даные из подгружаемой функции в файл
        if left == False:
            pass
        else:
            start_column = 1
            for i,item in enumerate(self.print_list[day],1):
                if i <=16:
                    self.worksheet.cell(start_row+10+i,start_column,i).font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+10+i,start_column).alignment = Alignment(horizontal="center",vertical='center')
                    self.worksheet.cell(start_row+10+i,start_column+1,item[1]).font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+10+i,start_column+1).alignment = Alignment(vertical='center')
                    self.worksheet.cell(start_row+10+i,start_column+2,item[2]).font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+10+i,start_column+2).alignment = Alignment(horizontal="center",vertical='center')
                    self.worksheet.cell(start_row+10+i,start_column+3,item[0]).font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+10+i,start_column+3).alignment = Alignment(horizontal="center",vertical='center')
                    self.worksheet.cell(start_row+10+i,start_column+4,"1").font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+10+i,start_column+4).alignment = Alignment(horizontal="center",vertical='center')
                else:
                    self.worksheet.cell(start_row+12+i,start_column+8,i).font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+12+i,start_column+8).alignment = Alignment(horizontal="center",vertical='center')
                    self.worksheet.cell(start_row+12+i,start_column+9,item[1]).font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+12+i,start_column+9).alignment = Alignment(vertical='center')
                    self.worksheet.cell(start_row+12+i,start_column+10,item[2]).font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+12+i,start_column+10).alignment = Alignment(horizontal="center",vertical='center')
                    self.worksheet.cell(start_row+12+i,start_column+11,item[0]).font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+12+i,start_column+11).alignment = Alignment(horizontal="center",vertical='center')
                    self.worksheet.cell(start_row+12+i,start_column+12,"1").font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+12+i,start_column+12).alignment = Alignment(horizontal="center",vertical='center')

        if right == False:
            pass
        else:
             for i,item in enumerate(self.print_list[day+1],1):
                if i <=16:
                    self.worksheet.cell(start_row+10+i,start_column+8,i).font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+10+i,start_column+8).alignment = Alignment(horizontal="center",vertical='center')
                    self.worksheet.cell(start_row+10+i,start_column+9,item[1]).font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+10+i,start_column+9).alignment = Alignment(vertical='center')
                    self.worksheet.cell(start_row+10+i,start_column+10,item[2]).font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+10+i,start_column+10).alignment = Alignment(horizontal="center",vertical='center')
                    self.worksheet.cell(start_row+10+i,start_column+11,item[0]).font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+10+i,start_column+11).alignment = Alignment(horizontal="center",vertical='center')
                    self.worksheet.cell(start_row+10+i,start_column+12,"1").font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+10+i,start_column+12).alignment = Alignment(horizontal="center",vertical='center')

                else:
                    self.worksheet.cell(start_row+12+i,start_column,i).font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+12+i,start_column).alignment = Alignment(horizontal="center",vertical='center')
                    self.worksheet.cell(start_row+12+i,start_column+1,item[1]).font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+12+i,start_column+1).alignment = Alignment(vertical='center')
                    self.worksheet.cell(start_row+12+i,start_column+2,item[2]).font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+12+i,start_column+2).alignment = Alignment(horizontal="center",vertical='center')
                    self.worksheet.cell(start_row+12+i,start_column+3,item[0]).font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+12+i,start_column+3).alignment = Alignment(horizontal="center",vertical='center')
                    self.worksheet.cell(start_row+12+i,start_column+4,"1").font = Font(size = 11,name = 'Times New Roman')
                    self.worksheet.cell(start_row+12+i,start_column+4).alignment = Alignment(horizontal="center",vertical='center')
    def save_workbook(self):
        if not os.path.exists(f"Отчеты по талонам\\{self.place}\\{self.Month_function(self.month)}"):
            os.makedirs(f"Отчеты по талонам\\{self.place}\\{self.Month_function(self.month)}")
        path = f"Отчеты по талонам\\{self.place}\\{self.Month_function(self.month)}\\{self.start_data}-{self.final_data}.xlsx"
        self.workbook.save(path)
        
        xlApp = Dispatch('Excel.Application')
        path = os.path.abspath (f'Отчеты по талонам\\{self.place}\\{self.Month_function(self.month)}\\{self.start_data}-{self.final_data}.xlsx')
        wb = xlApp.Workbooks.open(path)
        wb.SaveAs(path[:-1], FileFormat=56)
        xlApp.Quit()
        os.remove(f"{path}")
if __name__=="__main__":
    pr = PRINT_TO_XLS(start_data=1,final_data=8,month=4)
    pr.main()
    